from django.shortcuts import render
from django.db import models
from django.http import HttpResponse
import openpyxl
from raw_material.models import RmMaterialIssue, RmMaterialIssueSub, RawInwardMaterial, RawInwardMaterialSub
from packing_materials.models import PackingInwardMaterial, PackingInwardMaterialSub



def rm_issuance_report(request):
    """Fetch RM Issuance data dynamically based on filters, but hide initially."""
    issue_date_from = request.GET.get("from_date")
    issue_date_to = request.GET.get("to_date")
    search_query = request.GET.get("search")  # Search by item_code or issued_to_whom

    rm_issuances = RmMaterialIssueSub.objects.select_related("issue").none()  # Hide data initially

    if issue_date_from and issue_date_to:
        rm_issuances = RmMaterialIssueSub.objects.select_related("issue").filter(
            issue__date_of_issue__range=[issue_date_from, issue_date_to]
        )
        if search_query:
            rm_issuances = rm_issuances.filter(
                models.Q(item_code__icontains=search_query) |
                models.Q(issue__issued_to_whom__icontains=search_query)
            )
    for issuance in rm_issuances:
        item_details = ItemDetail.objects.filter(id=issuance.item_code).first()
        if item_details:
            issuance.item_name = item_details.item_name
            issuance.item_code = item_details.item_code# ✅ Correcting item_name

    no_data_message = "No records found for the selected filters." if issue_date_from and issue_date_to and not rm_issuances.exists() else ""

    return render(request, "reports/rm_issuance_report.html", {
        "rm_issuances": rm_issuances,
        "no_data_message": no_data_message
    })


def export_rm_issuance_excel(request):
    """Export RM Issuance Report to Excel (Customer-Wise)."""
    issue_date_from = request.GET.get("from_date")
    issue_date_to = request.GET.get("to_date")
    search_query = request.GET.get("search")

    # Fetch RM issuance data from RmMaterialIssueSub
    rm_issuances = RmMaterialIssueSub.objects.select_related("issue").all()

    if issue_date_from and issue_date_to:
        rm_issuances = rm_issuances.filter(issue__date_of_issue__range=[issue_date_from, issue_date_to])

    if search_query:
        rm_issuances = rm_issuances.filter(
            models.Q(item_code__icontains=search_query) |
            models.Q(issue__issued_to_whom__icontains=search_query)
        )

    # Check if there's data to export
    if not rm_issuances.exists():
        return HttpResponse("No data available for export.", content_type="text/plain")

    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RM Issuance Report"

    headers = ["Issue No", "Issue Date", "Issued To", "Item Code", "Item Name", "UOM", "Quantity", "Bags Issued", "Batch No"]
    ws.append(headers)

    for issuance in rm_issuances:
        ws.append([
            issuance.issue.iss_no,
            issuance.issue.date_of_issue.strftime("%Y-%m-%d"),
            issuance.issue.issued_to_whom,
            issuance.item_code,
            issuance.item_name,
            issuance.uom,
            issuance.quantity,
            issuance.bags_issued,
            issuance.batch_no,
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="rm_issuance_report.xlsx"'
    wb.save(response)

    return response


def generate_report(request):
    """Fetch report data dynamically based on filters, but hide initially."""
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    search_query = request.GET.get("search")  # Search by item_code or item_name

    raw_queryset = RawInwardMaterial.objects.none()  # Hide data initially
    packing_queryset = PackingInwardMaterial.objects.none()  # Hide data initially

    if from_date and to_date:
        raw_queryset = RawInwardMaterial.objects.filter(
            grn_date__range=[from_date, to_date]
        ).prefetch_related('rawinwardmaterialsub_set')
        packing_queryset = PackingInwardMaterial.objects.filter(
            grn_date__range=[from_date, to_date]
        ).prefetch_related('packinginwardmaterialsub_set')
        if search_query:
            raw_queryset = raw_queryset.filter(
                models.Q(vendor_code__icontains=search_query) |
                models.Q(vendor_name__icontains=search_query) |
                models.Q(rawinwardmaterialsub__item_code__icontains=search_query) |
                models.Q(rawinwardmaterialsub__item_name__icontains=search_query)
            )
            packing_queryset = packing_queryset.filter(
                models.Q(vendor_code__icontains=search_query) |
                models.Q(vendor_name__icontains=search_query) |
                models.Q(packinginwardmaterialsub__item_code__icontains=search_query) |
                models.Q(packinginwardmaterialsub__item_name__icontains=search_query)
            )

    results = []
    for raw in raw_queryset:
        if raw.rawinwardmaterialsub_set.exists():
            for raw_sub in raw.rawinwardmaterialsub_set.all():
                results.append({
                    'grn_no': raw.grn_no,
                    'grn_date': raw.grn_date,
                    'invoice_no': raw.invoice_no,
                    'vendor_code': raw.vendor_code,
                    'vendor_name': raw.vendor_name,
                    'item_code': raw_sub.item_code,
                    'item_name': raw_sub.item_name,
                    'quantity': raw_sub.quantity,
                    'uom': raw_sub.uom,
                    'batch_no': raw_sub.lot_no  # Use lot_no for batch number
                })
        else:
            results.append({
                'grn_no': raw.grn_no,
                'grn_date': raw.grn_date,
                'invoice_no': raw.invoice_no,
                'vendor_code': raw.vendor_code,
                'vendor_name': raw.vendor_name,
                'item_code': '',
                'item_name': '',
                'quantity': '',
                'uom': '',
                'batch_no': ''
            })
    for packing in packing_queryset:
        if packing.packinginwardmaterialsub_set.exists():
            for packing_sub in packing.packinginwardmaterialsub_set.all():
                results.append({
                    'grn_no': packing.grn_no,
                    'grn_date': packing.grn_date,
                    'invoice_no': packing.invoice_no,
                    'vendor_code': packing.vendor_code,
                    'vendor_name': packing.vendor_name,
                    'item_code': packing_sub.item_code,
                    'item_name': packing_sub.item_name,
                    'quantity': packing_sub.quantity,
                    'uom': packing_sub.uom,
                    # 'batch_no': packing_sub.batch_no  # Use batch_no for batch number
                })
        else:
            results.append({
                'grn_no': packing.grn_no,
                'grn_date': packing.grn_date,
                'invoice_no': packing.invoice_no,
                'vendor_code': packing.vendor_code,
                'vendor_name': packing.vendor_name,
                'item_code': '',
                'item_name': '',
                'quantity': '',
                'uom': '',
                'batch_no': ''
            })

    no_data_message = "No records found for the selected filters." if from_date and to_date and not results else ""

    return render(request, "reports/report.html", {
        "results": results,
        "no_data_message": no_data_message,
    })

def export_report_excel(request):
    """Export report data to Excel."""
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    search_query = request.GET.get("search")

    raw_queryset = RawInwardMaterial.objects.all().prefetch_related('rawinwardmaterialsub_set')
    packing_queryset = PackingInwardMaterial.objects.all().prefetch_related('packinginwardmaterialsub_set')

    if from_date and to_date:
        raw_queryset = raw_queryset.filter(grn_date__range=[from_date, to_date])
        packing_queryset = packing_queryset.filter(grn_date__range=[from_date, to_date])

    if search_query:
        raw_queryset = raw_queryset.filter(
            models.Q(vendor_code__icontains=search_query) |
            models.Q(vendor_name__icontains=search_query) |
            models.Q(rawinwardmaterialsub__item_code__icontains=search_query) |
            models.Q(rawinwardmaterialsub__item_name__icontains=search_query)
        )
        packing_queryset = packing_queryset.filter(
            models.Q(vendor_code__icontains=search_query) |
            models.Q(vendor_name__icontains=search_query) |
            models.Q(packinginwardmaterialsub__item_code__icontains=search_query) |
            models.Q(packinginwardmaterialsub__item_name__icontains=search_query)
        )

    results = []
    for raw in raw_queryset:
        if raw.rawinwardmaterialsub_set.exists():
            for raw_sub in raw.rawinwardmaterialsub_set.all():
                results.append({
                    'grn_no': raw.grn_no,
                    'grn_date': raw.grn_date,
                    'invoice_no': raw.invoice_no,
                    'vendor_code': raw.vendor_code,
                    'vendor_name': raw.vendor_name,
                    'item_code': raw_sub.item_code,
                    'item_name': raw_sub.item_name,
                    'quantity': raw_sub.quantity,
                    'uom': raw_sub.uom,
                    'batch_no': raw_sub.repacking_batch_no  # Use lot_no for batch number
                })
        else:
            results.append({
                'grn_no': raw.grn_no,
                'grn_date': raw.grn_date,
                'invoice_no': raw.invoice_no,
                'vendor_code': raw.vendor_code,
                'vendor_name': raw.vendor_name,
                'item_code': '',
                'item_name': '',
                'quantity': '',
                'uom': '',
                'batch_no': ''
            })
    for packing in packing_queryset:
        if packing.packinginwardmaterialsub_set.exists():
            for packing_sub in packing.packinginwardmaterialsub_set.all():
                results.append({
                    'grn_no': packing.grn_no,
                    'grn_date': packing.grn_date,
                    'invoice_no': packing.invoice_no,
                    'vendor_code': packing.vendor_code,
                    'vendor_name': packing.vendor_name,
                    'item_code': packing_sub.item_code,
                    'item_name': packing_sub.item_name,
                    'quantity': packing_sub.quantity,
                    'uom': packing_sub.uom,
                    # 'batch_no': packing_sub.batch_no  # Use batch_no for batch number
                })
        else:
            results.append({
                'grn_no': packing.grn_no,
                'grn_date': packing.grn_date,
                'invoice_no': packing.invoice_no,
                'vendor_code': packing.vendor_code,
                'vendor_name': packing.vendor_name,
                'item_code': '',
                'item_name': '',
                'quantity': '',
                'uom': '',
                'batch_no': ''
            })

    if not results:
        return HttpResponse("No data available for export.", content_type="text/plain")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = ["GRN No", "GRN Date", "Inv No", "Vendor Code", "Vendor Name", "Item Code", "Item Name", "Qty", "UOM", "Batch No"]
    ws.append(headers)

    for item in results:
        ws.append([
            item['grn_no'],
            item['grn_date'].strftime("%Y-%m-%d"),
            item['invoice_no'],
            item['vendor_code'],
            item['vendor_name'],
            item['item_code'],
            item['item_name'],
            item['quantity'],
            # item['uom'],
            # item['batch_no']
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="report.xlsx"'
    wb.save(response)

    return response





from django.shortcuts import render
from django.db import models
from django.http import HttpResponse
import openpyxl
from packing_materials.models import pmmaterialissue, pmmaterialissuesub,PackingInwardMaterial
from master.models import ItemDetail  # ✅ Corrected Import for Vendor Master Table

def pm_material_issue_report(request):
    """Fetch PM Material Issuance data dynamically based on filters, but hide initially."""
    issue_date_from = request.GET.get("from_date")
    issue_date_to = request.GET.get("to_date")
    search_query = request.GET.get("search")  # Search by item_code or issued_to_whom

    rm_issuances = pmmaterialissuesub.objects.select_related("matIssueId").none()  # 🔹 Hide data initially

    if issue_date_from and issue_date_to:
        rm_issuances = pmmaterialissuesub.objects.select_related("matIssueId").filter(
            matIssueId__issue_date__range=[issue_date_from, issue_date_to]
        )
        if search_query:
            rm_issuances = rm_issuances.filter(
                models.Q(item_code__icontains=search_query) |
                models.Q(matIssueId__issue_to_whom__icontains=search_query)
            )
    for issuance in rm_issuances:
        item_details = ItemDetail.objects.filter(id=issuance.item_code).first()
        if item_details:
            issuance.item_name = item_details.item_name
            issuance.item_code = item_details.item_code# ✅ Correcting item_name

    no_data_message = "No records found for the selected filters." if issue_date_from and issue_date_to and not rm_issuances.exists() else ""

    return render(request, "reports/pm_material_issue_report.html", {
        "rm_issuances": rm_issuances,
        "no_data_message": no_data_message
    })


def export_rm_issuance_excel(request):
    """Export PM Material Issuance Report to Excel."""
    issue_date_from = request.GET.get("from_date")
    issue_date_to = request.GET.get("to_date")
    search_query = request.GET.get("search")

    # Fetch PM issuance data
    rm_issuances = pmmaterialissuesub.objects.select_related("matIssueId").all()

    if issue_date_from and issue_date_to:
        rm_issuances = rm_issuances.filter(
            matIssueId__issue_date__range=[issue_date_from, issue_date_to]
        )

    if search_query:
        rm_issuances = rm_issuances.filter(
            models.Q(item_code__icontains=search_query) |
            models.Q(item_name__icontains=search_query)
        )

    # Check if there's data to export
    if not rm_issuances.exists():
        return HttpResponse("No data available for export.", content_type="text/plain")

    # Create Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PM Material Issue Report"

    headers = [ "Issue No", "Issue Date", "Item Code", "Item Name", "Quantity"]
    ws.append(headers)

    for issuance in rm_issuances:
        # Ensure matIssueId is not NULL
        if not issuance.matIssueId:
            continue  # Skip if there's no related matIssueId

        # ✅ Fetch Vendor Name from Master Table
        # vendor = PackingInwardMaterial.objects.filter(issue_id=issuance.matIssueId).first()
        # vendor_name = vendor.vendor if vendor else "Unknown"


        # Debugging print statements
        print(f"Issue No: {issuance.matIssueId.issue_no}")
        print(f"Issue Date: {issuance.matIssueId.issue_date}")
        

        ws.append([
           
            issuance.matIssueId.issue_no,
            issuance.matIssueId.issue_date.strftime("%Y-%m-%d"),
            issuance.item_code,
            issuance.item_name,
            issuance.quantity,
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="pm_material_issue_report.xlsx"'
    wb.save(response)

    return response




from django.shortcuts import render
from django.db.models import Q
from django.http import HttpResponse
import openpyxl
from finished_goods.models import PackingSlip, PackingSlipItem

def packing_slip_report(request):
    """Fetch Packing Slip data dynamically based on filters, but hide initially."""
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    search_query = request.GET.get("search")  # Search by customer name or item code

    packing_slips = PackingSlip.objects.none()  # Hide data initially

    if from_date and to_date:
        packing_slips = PackingSlip.objects.filter(
            ps_date__range=[from_date, to_date]
        ).prefetch_related('items')

        if search_query:
            packing_slips = packing_slips.filter(
                Q(customer__customer_name__icontains=search_query) |
                Q(items__item_code__item_code__icontains=search_query)
            ).distinct()

    no_data_message = "No records found for the selected filters." if from_date and to_date and not packing_slips.exists() else ""

    return render(request, "reports/packing_slip_report.html", {
        "packing_slips": packing_slips,
        "no_data_message": no_data_message
    })

def export_packing_slip_excel(request):
    """Export Packing Slip Report to Excel."""
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    search_query = request.GET.get("search")

    packing_slips = PackingSlip.objects.all().prefetch_related('items')

    if from_date and to_date:
        packing_slips = packing_slips.filter(ps_date__range=[from_date, to_date])

    if search_query:
        packing_slips = packing_slips.filter(
            Q(customer__customer_name__icontains=search_query) |
            Q(items__item_code__item_code__icontains=search_query)
        ).distinct()

    if not packing_slips.exists():
        return HttpResponse("No data available for export.", content_type="text/plain")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packing Slip Report"

    headers = ["Customer Name", "PS No", "PS Date", "Item Code", "Item Name", "No of Bags", "Qty", "Batch No"]
    ws.append(headers)

    for slip in packing_slips:
        for item in slip.items.all():
            ws.append([
                slip.customer.customer_name,
                slip.ps_no,
                slip.ps_date.strftime("%Y-%m-%d"),
                item.item_code.item_code,
                item.item_name,
                item.box_bags,
                item.qty,
                item.batch_no
            ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="packing_slip_report.xlsx"'
    wb.save(response)

    return response



from django.shortcuts import render
from django.db.models import Q
from django.http import HttpResponse
import openpyxl
from finished_goods.models import FGInwardMaterial, FGInwardMaterialSub

def fg_inward_material_report(request):
    """Fetch FG Inward Material data dynamically based on filters, but hide initially."""
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    search_query = request.GET.get("search")  # Search by item code or item name

    inward_materials = FGInwardMaterial.objects.none()  # Hide data initially

    if from_date and to_date:
        inward_materials = FGInwardMaterial.objects.filter(
            inward_date__range=[from_date, to_date]
        ).prefetch_related('fginwardmaterialsub_set')

        if search_query:
            inward_materials = inward_materials.filter(
                Q(fginwardmaterialsub__item_code__icontains=search_query) |
                Q(fginwardmaterialsub__item_name__icontains=search_query)
            ).distinct()

    no_data_message = "No records found for the selected filters." if from_date and to_date and not inward_materials.exists() else ""

    return render(request, "reports/fg_inward_material_report.html", {
        "inward_materials": inward_materials,
        "no_data_message": no_data_message
    })

def export_fg_inward_material_excel(request):
    """Export FG Inward Material Report to Excel."""
    from_date = request.GET.get("from_date")
    to_date = request.GET.get("to_date")
    search_query = request.GET.get("search")

    inward_materials = FGInwardMaterial.objects.all().prefetch_related('fginwardmaterialsub_set')

    if from_date and to_date:
        inward_materials = inward_materials.filter(inward_date__range=[from_date, to_date])

    if search_query:
        inward_materials = inward_materials.filter(
            Q(fginwardmaterialsub__item_code__icontains=search_query) |
            Q(fginwardmaterialsub__item_name__icontains=search_query)
        ).distinct()

    if not inward_materials.exists():
        return HttpResponse("No data available for export.", content_type="text/plain")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FG Inward Material Report"

    headers = ["Inv No", "Inv Date", "Item Code", "Item Name", "No of Bags", "Qty", "UOM", "Batch No"]
    ws.append(headers)

    for material in inward_materials:
        for item in material.fginwardmaterialsub_set.all():
            ws.append([
                material.inward_no,
                material.inward_date.strftime("%Y-%m-%d"),
                item.item_code,
                item.item_name,
                item.box_no,
                item.quantity,
                item.uom,
                item.batch_no
            ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="fg_inward_material_report.xlsx"'
    wb.save(response)

    return response

from django.shortcuts import render
from .models import StockStatement

def stock_dashboard(request):
    selected_category = request.GET.get('category', 'ALL')

    if selected_category == 'ALL':
        stocks = StockStatement.objects.all()
    else:
        stocks = StockStatement.objects.filter(category=selected_category)

    context = {'stocks': stocks}
    return render(request, 'reports/stock_dashboard.html', context)




